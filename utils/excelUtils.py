import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import os
import numpy as np

class excelUtils:
    # Función para limpiar el archivo Excel
    def cleanExcel(self, filePath):
        # Leer el archivo Excel
        df = pd.read_excel(filePath, dtype=str)

        # Eliminar filas que contienen "Total" o "Filtros aplicados"
        df_cleaned = df[~df.iloc[:, 0].str.contains('Total|Filtros aplicados', na=False)].copy()


        return df_cleaned
    


    def filterByPeriod(self, df, period):
        """
        Filtrar las filas que coincidan con el período proporcionado en formato 'YYYY-MM'.
        Cualquier registro fuera del período será eliminado.
        """
        try:
            # Verificar si la columna 'Fecha' existe
            if 'Fecha' not in df.columns:
                raise ValueError("No se encontró la columna 'Fecha' en el archivo.")

            # Convertir la columna 'Fecha' a tipo datetime usando el formato adecuado
            df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce')

            # Extraer el período en formato 'YYYY-MM'
            df['Periodo'] = df['Fecha'].dt.strftime('%Y-%m')

            # Filtrar las filas que coincidan con el período proporcionado
            df_filtered = df[df['Periodo'] == period].copy()

            # Eliminar la columna 'Periodo' ya que no es necesaria
            df_filtered.drop(columns=['Periodo'], inplace=True)

            # Formatear la columna 'Fecha' de nuevo a 'DD/MM/YYYY' para conservar el formato original
            df_filtered['Fecha'] = df_filtered['Fecha'].dt.strftime('%d/%m/%Y')

            # Retornar el dataframe filtrado
            return df_filtered

        except Exception as e:
            print(f"Error al filtrar por período: {str(e)}")
            raise


    # Función para guardar los datos limpiados
    def saveCleanedData(self, df_cleaned, filePath):
        df_cleaned.to_excel(filePath, index=False, engine='openpyxl')

    # Función para aplicar formato "Texto" a todas las celdas
    def applyTextFormat(self, filePath):
        # Cargar el archivo con openpyxl
        wb = load_workbook(filePath)
        ws = wb.active

        # Crear estilo de texto
        text_style = NamedStyle(name="text_style", number_format="@")  # Formato "Texto"

        # Aplicar el formato a todas las celdas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.style = text_style

        # Guardar el archivo con el formato aplicado
        wb.save(filePath)

    # Generar Archivo para Acepta
    def process_excel_file(self, filePath):
        # Leer el archivo Excel y procesar las columnas
        df = pd.read_excel(filePath, dtype=str)
        df['Serie1-Serie2'] = df['Serie1'] + '-' + df['Serie2']
        df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce').dt.strftime('%Y-%m-%d')
        df_acepta = df[['TipoDocumento', 'Serie1-Serie2', 'Fecha']].copy()
        df_acepta.loc[:, 'Error'] = 'Error en Sistema'
        return df_acepta


    def get_output_directory(self, filePath):
        # Obtener la ruta de la carpeta 'acepta' (ya existente)
        output_dir = os.path.join(os.path.dirname(filePath), 'acepta')
        return output_dir
    

    def split_and_save_csv_by_day(self, df_acepta, output_dir, rows_per_file=45):
        try:
            # Eliminar espacios en blanco o caracteres erróneos en las fechas
            df_acepta['Fecha'] = df_acepta['Fecha'].str.strip()

            # Convertir las fechas con el formato correcto: yyyy-mm-dd
            df_acepta['Fecha'] = pd.to_datetime(df_acepta['Fecha'], format='%Y-%m-%d', errors='coerce')

            # Verificar si hubo algún error al convertir las fechas
            if df_acepta['Fecha'].isnull().any():
                raise ValueError("Algunas fechas no se pudieron convertir correctamente.")

            # Agrupar por la columna 'Fecha'
            grouped = df_acepta.groupby('Fecha')

            # Para llevar la cuenta de cuántos archivos se crean en total
            total_files_created = 0

            # Para cada grupo de fecha, dividir y guardar en archivos CSV de 45 filas
            for fecha, group in grouped:
                num_parts = (len(group) // rows_per_file) + (1 if len(group) % rows_per_file != 0 else 0)
                for i in range(num_parts):
                    part_df = group.iloc[i * rows_per_file:(i + 1) * rows_per_file]
                    # Crear el nombre del archivo basado en la fecha y el índice de la parte
                    formatted_date = fecha.strftime('%Y-%m-%d')
                    output_path = os.path.join(output_dir, f'acepta_{formatted_date}_parte_{i + 1}.csv')
                    part_df.to_csv(output_path, sep=';', index=False, header=False)
                    total_files_created += 1

            return total_files_created

        except FileNotFoundError as e:
            return f"Error: Archivo o directorio no encontrado - {str(e)}"
        
        except ValueError as e:
            return f"Error: {str(e)}"
        
        except Exception as e:
            return f"Error inesperado: {str(e)}"


    
    # Generar archivo para sunat
    def process_excel_file_sunat(self, filePath):
        df = pd.read_excel(filePath, dtype=str)
        df_sunat = df[['Empresa', 'TipoDocumento', 'Serie1', 'Serie2', 'Fecha', 'Importe Total']].copy()
        return df_sunat

    def get_output_directory_sunat(self, filePath):
        return os.path.join(os.path.dirname(filePath), 'sunat')

    def split_and_save_txt(self, df_sunat, output_dir, rows_per_file=100):
        # Eliminar filas que estén completamente vacías o tengan solo espacios
        df_sunat = df_sunat.dropna(how='all').apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        df_sunat = df_sunat.replace('', np.nan).dropna(how='any')
        
        num_parts = (len(df_sunat) // rows_per_file) + (1 if len(df_sunat) % rows_per_file != 0 else 0)
        
        for i in range(num_parts):
            part_df = df_sunat.iloc[i * rows_per_file:(i + 1) * rows_per_file]

            # Convertir el DataFrame en una lista de líneas de texto
            output_path = os.path.join(output_dir, f'sunat_{i + 1}.txt')
            
            # Guardar el DataFrame en un archivo temporal antes de aplicar la limpieza
            part_df.to_csv(output_path, sep='|', index=False, header=False)
            
            # Leer el archivo temporal para limpiar líneas vacías o con espacios
            with open(output_path, 'r') as file:
                lines = file.readlines()

            # Eliminar líneas que están vacías o que solo contienen espacios
            lines = [line.strip() for line in lines if line.strip()]

            # Guardar el archivo sin líneas vacías al final
            with open(output_path, 'w') as file:
                file.write('\n'.join(lines))
            
        return num_parts