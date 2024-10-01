import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

class LowModel:

    # Limpiar el archivo Excel y aplicar formato de texto a todas las columnas
    def fileCleanup(self, filePath):
        try:
            # Leer el archivo Excel con pandas, procesar como texto para evitar errores
            df = pd.read_excel(filePath, dtype=str)

            # Identificar y eliminar las filas que contienen "Total" y "Filtros aplicados"
            df_cleaned = df[~df.iloc[:, 0].str.contains('Total|Filtros aplicados', na=False)].copy()

            # Formatear la columna 'Fecha' para mostrar en formato 'dd/mm/yyyy'
            if 'Fecha' in df_cleaned.columns:
                # Convertir a datetime, luego a string con formato día/mes/año
                df_cleaned['Fecha'] = pd.to_datetime(df_cleaned['Fecha'], errors='coerce').dt.strftime('%d/%m/%Y')

            # Guardar los datos limpiados en el mismo archivo (sobrescribir)
            df_cleaned.to_excel(filePath, index=False, engine='openpyxl')

            # Cargar el archivo con openpyxl para cambiar el formato de todas las celdas a "Texto"
            wb = load_workbook(filePath)
            ws = wb.active

            # Definir el estilo de texto
            text_style = NamedStyle(name="text_style", number_format="@")  # Formato de texto ("@")

            # Aplicar formato "Texto" a todas las celdas
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.style = text_style

            # Guardar el archivo con el formato correcto
            wb.save(filePath)

            print(f"Archivo {filePath} limpiado y formateado correctamente.")
            return {'success': True, 'message': 'Archivo limpiado y formateado correctamente.'}

        except Exception as e:
            print(f"Error al limpiar y formatear el archivo: {str(e)}")
            return {'success': False, 'message': f'Error al limpiar y formatear el archivo: {str(e)}'}



    # crear archivo acepta
    def createFileAcepta(self, urlFile):
        pass

    # crear archivo sunat
    def createFileSunat(self, urlFile):
        pass